"""
SkillBridge Extractor - Any Job Family
--------------------------------------------------------------
✅ Extracts all SkillBridge listings for one job family via API
✅ Handles all 300+ pages automatically
✅ Retries failed pages
✅ Exports clean Excel
✅ Shows a live tqdm progress bar
"""

import requests
import pandas as pd
import time
import os
from datetime import datetime
from tqdm import tqdm
from openpyxl.styles import Font, Alignment, PatternFill

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
BASE_URL = "https://sb-api.azurewebsites.us/Location/Lookup"
PAGE_SIZE = 10
MAX_RETRIES = 3

OUTPUT_DIR = os.path.join(
    r"C:\Users\danny\OneDrive\Desktop\skillbridge\Skillbridge-Web-Scraper correction",
    "output"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ------------------------------------------------------------
# FUNCTION: Fetch all pages for one job family
# ------------------------------------------------------------
def fetch_job_family_data(industry):
    print(f"📦 Extracting SkillBridge data for: {industry}\n")

    all_rows = []
    start = 0
    total_records = None
    total_pages = None

    # first request to detect total pages
    initial_params = {
        "draw": 1,
        "order[0][column]": 19,
        "order[0][dir]": "asc",
        "start": 0,
        "length": PAGE_SIZE,
        "search[value]": "",
        "industry": industry,
        "colMatrix": "1-Organization,2-Program,3-Branch,4-City,5-State,6-Zip,7-Duration,8-EmployerPoc,9-EmployerPocEmail,10-DeliveryMethodId,11-Installation,12-LocationStates,13-TargetMOCs,14-OtherEligibilityFactors,15-Other,16-JobDescription,17-Summary,18-Industries,19-Distance",
        "device": "platform:Windows,browser:Chrome",
        "mobile": "false"
    }

    first_response = requests.get(BASE_URL, params=initial_params, timeout=30)
    first_data = first_response.json()
    total_records = first_data.get("recordsTotal", 0)
    total_pages = (total_records // PAGE_SIZE) + 1
    print(f"🧾 Total records: {total_records} → {total_pages} pages\n")

    # fetch all pages with tqdm progress bar
    for page in tqdm(range(total_pages), desc=f"{industry} pages", ncols=90):
        start = page * PAGE_SIZE
        params = initial_params.copy()
        params["start"] = start

        for attempt in range(MAX_RETRIES):
            try:
                response = requests.get(BASE_URL, params=params, timeout=30)
                response.raise_for_status()
                json_data = response.json()
                data = json_data.get("data", [])
                if data:
                    all_rows.extend(data)
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2)
                else:
                    print(f"\n❌ Failed page {page + 1}: {e}")

        time.sleep(0.3)

    return all_rows


# ------------------------------------------------------------
# FUNCTION: Export to Excel
# ------------------------------------------------------------
def export_to_excel(data, industry):
    if not data:
        print("❌ No data found — nothing to export.")
        return

    df = pd.DataFrame(data)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"SkillBridge_{industry.replace(' ', '_')}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SkillBridge", index=False)
        ws = writer.sheets["SkillBridge"]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.freeze_panes = "A2"

    print(f"\n✅ Exported {len(df)} records → {filepath}")


# ------------------------------------------------------------
# MAIN EXECUTION
# ------------------------------------------------------------
if __name__ == "__main__":
    print("🚀 Starting SkillBridge extraction...\n")

    job_family = input("Enter Job Family exactly as shown on the SkillBridge site: ").strip()
    if not job_family:
        print("❌ No job family entered. Exiting.")
        exit()

    data = fetch_job_family_data(job_family)
    export_to_excel(data, job_family)

    print("\n🎯 Done! All data saved successfully.\n")
