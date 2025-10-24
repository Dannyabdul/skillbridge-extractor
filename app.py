from flask import Flask, request, jsonify
import requests
import pandas as pd
import time
import os
from datetime import datetime
from tqdm import tqdm
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)

BASE_URL = "https://sb-api.azurewebsites.us/Location/Lookup"
PAGE_SIZE = 10
MAX_RETRIES = 3
OUTPUT_DIR = os.path.join(r"C:\Users\danny\OneDrive\Desktop\skillbridge", "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ------------------------------------------------------------
# MAIN EXTRACTION FUNCTION
# ------------------------------------------------------------
def fetch_job_family_data(industry):
    print(f"📦 Extracting SkillBridge data for: {industry}\n")
    all_rows = []
    start = 0
    total_records = None
    total_pages = None

    initial_params = {
        "draw": 1,
        "order[0][column]": 19,
        "order[0][dir]": "asc",
        "start": 0,
        "length": PAGE_SIZE,
        "search[value]": "",
        "industry": industry,
        "colMatrix": "1-Organization,2-Program,3-Branch,4-City,5-State,6-Zip,7-Duration,8-EmployerPoc,9-EmployerPocEmail,10-DeliveryMethodId,11-Installation,12-LocationStates,13-TargetMOCs,14-OtherEligibilityFactors,15-Other,16-JobDescription,17-Summary,18-Industries,19-Distance",
        "device": "platform:Windows,browser:Chrome",
        "mobile": "false"
    }

    first_response = requests.get(BASE_URL, params=initial_params, timeout=30)
    first_data = first_response.json()
    total_records = first_data.get("recordsTotal", 0)
    total_pages = (total_records // PAGE_SIZE) + 1
    print(f"🧾 Total records: {total_records} → {total_pages} pages\n")

    for page in tqdm(range(total_pages), desc=f"{industry} pages", ncols=90):
        start = page * PAGE_SIZE
        params = initial_params.copy()
        params["start"] = start

        for attempt in range(MAX_RETRIES):
            try:
                response = requests.get(BASE_URL, params=params, timeout=30)
                response.raise_for_status()
                json_data = response.json()
                data = json_data.get("data", [])
                if data:
                    all_rows.extend(data)
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2)
                else:
                    print(f"\n❌ Failed page {page + 1}: {e}")
        time.sleep(0.3)

    return all_rows


def export_to_excel(data, industry):
    if not data:
        print("❌ No data found — nothing to export.")
        return None

    df = pd.DataFrame(data)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"SkillBridge_{industry.replace(' ', '_')}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SkillBridge", index=False)
        ws = writer.sheets["SkillBridge"]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.freeze_panes = "A2"

    print(f"\n✅ Exported {len(df)} records → {filepath}")
    return filepath


@app.route("/extract", methods=["GET"])
def extract_data():
    job_family = request.args.get("jobFamily", "Business and Financial Operations")
    data = fetch_job_family_data(job_family)
    output_path = export_to_excel(data, job_family)
    return jsonify({
        "status": "success",
        "records": len(data),
        "jobFamily": job_family,
        "output": output_path
    })


if __name__ == "__main__":
    # Automatically start an ngrok tunnel
    print("🌐 SkillBridge Flask app is running!")
    print(f"👉 Local:  http://127.0.0.1:5000/extract?jobFamily=Business%20and%20Financial%20Operations")
    print(f"🌎 Public: {public_url}/extract?jobFamily=Business%20and%20Financial%20Operations\n")
    app.run(host="0.0.0.0", port=5000)

